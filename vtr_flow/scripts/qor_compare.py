#!/usr/bin/env python3
import argparse
from collections import OrderedDict

import openpyxl  # Excel spreadsheet manipulation library
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.cell.cell import TYPE_NUMERIC
import os
import pandas as pd
import re

DEFAULT_METRICS = [
    # ABC QoR Metrics
    "abc_depth",
    "num_pre_packed_blocks",
    # Pack QoR Metrics
    "num_post_packed_blocks",
    "num_clb",  # VTR
    "num_memories",  # VTR
    "num_mult",  # VTR
    "num_LAB",  # Titan
    "num_DSP",  # Titan
    "num_M9K",  # Titan
    "num_M144K",  # Titan
    "device_grid_tiles",
    # Place QoR Metrics
    "placed_wirelength_est",
    "placed_CPD_est",
    # Route QoR Metrics
    "min_chan_width",  # VTR
    "routed_wirelength",  # VTR (minW), Titan
    "crit_path_routed_wirelength",  # VTR
    "critical_path_delay",  # VTR/Titan
    "geomean_nonvirtual_intradomain_critical_path_delay",  # VTR/Titan
    # Run-time Metrics
    "odin_synth_time",
    "abc_synth_time",
    "pack_time",
    "place_time",
    "min_chan_width_route_time",
    "crit_path_route_time",
    "vtr_flow_elapsed_time",
    "max_vpr_mem",
]

DEFAULT_KEYS = [
    "arch",
    "circuit",
]


def parse_args():

    parser = argparse.ArgumentParser(
        description="Utility script to generate spreadsheets comparing VTR metric files"
    )

    parser.add_argument(
        "parse_result_files",
        nargs="*",
        metavar="PARSE_RESULT.TXT",
        help="Set of metric files to compare",
    )

    parser.add_argument(
        "--result_names",
        nargs="+",
        metavar="RESULT_NAME",
        help="Name of for each metric file (default: derived from filenames)",
    )

    parser.add_argument(
        "-o",
        "--output_spreadsheet",
        metavar="OUTPUT.XLSX",
        default="comparison_output.xlsx",
        help="Output spreadsheet file (default: %(default)s)",
    )

    parser.add_argument(
        "--qor_metrics",
        nargs="+",
        default=DEFAULT_METRICS,
        help="QoR Metrics to compare (default: %(default)s)",
    )

    parser.add_argument(
        "--keys",
        nargs="+",
        default=DEFAULT_KEYS,
        help="Default keys to identify individual experiments (default: %(default)s)",
    )

    parser.add_argument(
        "-t",
        "--task",
        metavar="task_name",
        default="",
        help="Compare the latest result of a task to its golden results",
    )

    args = parser.parse_args()

    if args.task:
        args.parse_result_files = get_task_result_files(args.task)

    if not args.task and len(args.parse_result_files) == 0:
        raise RuntimeError("No parse result files or task name provided")

    if not args.result_names or len(args.result_names) == 0:
        args.result_names = args.parse_result_files
    elif len(args.result_names) != len(args.parse_result_files):
        raise RuntimeError(
            "Number of result names ({}) does not match number of result files ({})".format(
                len(args.result_names), len(args.parse_result_files)
            )
        )
    else:
        assert len(args.result_names) == len(args.parse_result_files)

    return args


def main():

    args = parse_args()

    wb = openpyxl.Workbook()

    wb.remove(wb.active)

    assert len(args.result_names) == len(args.parse_result_files)

    avail_metrics = set()

    # Load all the raw data into separate sheets
    raw_sheets = OrderedDict()
    for i, csv in enumerate(args.parse_result_files):
        # Load as CSV
        print("Loading", csv)

        base, ext = os.path.splitext(csv)
        if ext == ".txt":
            sep = "\t"
        else:
            sep = ","
        df = pd.read_csv(csv, sep=sep)

        avail_metrics.update(df.columns)  # Record available metrics

        # Convert to work sheet
        sheet_name = safe_sheet_title("{}".format(args.result_names[i]))
        ws = dataframe_to_sheet(wb, df, sheet_name)

        raw_sheets[sheet_name] = ws

    qor_metrics = []
    for metric in args.qor_metrics:
        if metric not in avail_metrics:
            print(
                "Warning: Metric",
                metric,
                "not found in parse results (may be for a different benchmark set)",
            )
        else:
            qor_metrics.append(metric)

    # Create sheet of ratios
    ratio_sheet = wb.create_sheet("ratios", index=0)
    ratio_ranges = make_ratios(ratio_sheet, raw_sheets, keys=args.keys, metrics=qor_metrics)

    # Create summary sheet
    summary_data_sheet = wb.create_sheet("summary_data", index=0)
    make_summary(summary_data_sheet, ratio_sheet, ratio_ranges, args.keys)

    summary_sheet = wb.create_sheet("summary", index=0)
    make_transpose(summary_sheet, summary_data_sheet)

    wb.save(args.output_spreadsheet)


def make_transpose(dest_sheet, ref_sheet):

    for ref_row in range(ref_sheet.min_row, ref_sheet.max_row + 1):
        for ref_col in range(ref_sheet.min_column, ref_sheet.max_column + 1):
            ref_cell = ref_sheet.cell(row=ref_row, column=ref_col)

            dest_cell = dest_sheet.cell(row=ref_col, column=ref_row)
            dest_cell.value = '=IF(OR(ISBLANK({cell}),ISERROR({cell})),"",{cell})'.format(
                cell=value_ref(ref_cell)
            )


def make_summary(summary_sheet, ratio_sheet, ratio_ranges, keys):
    dest_row = 1
    for i, (ratio_name, cell_range) in enumerate(ratio_ranges.items()):

        dest_col = 1

        if i == 0:  # First range, copy headers

            dest_col += 1
            ratio_row = cell_range.min_row
            for j, ratio_col in enumerate(range(cell_range.min_col, cell_range.max_col + 1)):
                ratio_cell = ratio_sheet.cell(row=ratio_row, column=ratio_col)

                if j < len(keys):
                    continue
                else:
                    dest_cell = summary_sheet.cell(row=dest_row, column=dest_col)
                    dest_cell.value = "={}".format(value_ref(ratio_cell))
                    dest_col += 1
            dest_row += 1
            dest_col = 1  # Reset for next row

        #
        # Data
        #

        # Ratio name
        dest_cell = summary_sheet.cell(row=dest_row, column=dest_col)
        dest_cell.value = ratio_name

        dest_col += 1

        # Ratio values
        ratio_row = cell_range.max_row  # Geomean is in last row
        for ratio_col in range(cell_range.min_col + len(keys), cell_range.max_col + 1):
            ratio_cell = ratio_sheet.cell(row=ratio_row, column=ratio_col)

            dest_cell = summary_sheet.cell(row=dest_row, column=dest_col)
            dest_cell.value = "={}".format(value_ref(ratio_cell))

            dest_col += 1

        dest_row += 1


def make_ratios(ratio_sheet, raw_sheets, keys, metrics):

    ref_sheet_title = list(raw_sheets.keys())[0]

    ref_sheet = raw_sheets[ref_sheet_title]  # Get the first raw sheet

    cell_ranges = OrderedDict()

    row = 1
    for raw_sheet_title, raw_sheet in raw_sheets.items():
        ratio_subtitle = ratio_sheet.cell(row=row, column=1)
        ratio_subtitle.value = raw_sheet_title
        row += 1

        cell_range = fill_ratio(
            ratio_sheet, raw_sheet, ref_sheet, dest_row=row, dest_col=1, keys=keys, metrics=metrics
        )
        row += cell_range.max_row - cell_range.min_row
        row += 2

        cell_ranges[raw_sheet_title] = cell_range

    return cell_ranges


def fill_ratio(ws, raw_sheet, ref_sheet, dest_row, dest_col, keys, metrics):

    cell_range = CellRange(
        min_row=dest_row, max_row=dest_row, min_col=dest_col, max_col=dest_col, title=ws.title
    )

    link_sheet_header(ws, ref_sheet, row=dest_row, values=keys + metrics)
    dest_row += 1

    col_offset = 0
    for ref_col in range(1, ref_sheet.max_column + 1):
        ref_header = ref_sheet.cell(row=1, column=ref_col)
        raw_header = raw_sheet.cell(row=1, column=ref_col)

        ref_header_name = ref_header.value.strip()
        raw_header_name = ref_header.value.strip()
        assert ref_header_name == raw_header_name, "Header values must match"

        if ref_header_name not in keys + metrics:
            continue

        row_offset = 0
        for ref_row in range(2, ref_sheet.max_row + 1):

            ref_cell = ref_sheet.cell(row=ref_row, column=ref_col)
            raw_cell = raw_sheet.cell(row=ref_row, column=ref_col)

            dest_cell = ws.cell(row=dest_row + row_offset, column=dest_col + col_offset)

            if ref_header_name in keys:
                assert ref_cell.value == raw_cell.value, "Key value must match"

                dest_cell.value = "={}".format(value_ref(ref_cell))
            elif ref_header.value in metrics and ref_cell.data_type == TYPE_NUMERIC:
                dest_cell.value = "={}".format(safe_ratio_ref(raw_cell, ref_cell))
            else:
                pass
                # assert False
            row_offset += 1

        # Last row: Add GEOMEAN
        dest_cell = ws.cell(row=dest_row + row_offset, column=dest_col + col_offset)
        if ref_header.value in keys:
            if keys.index(ref_header.value) == len(keys) - 1:
                # Last key make geomean title
                dest_cell.value = "GEOMEAN"
        else:
            start_cell = ws.cell(
                row=dest_row + row_offset - ref_sheet.max_row + 1, column=dest_col + col_offset
            )
            end_cell = ws.cell(row=dest_row + row_offset - 1, column=dest_col + col_offset)

            dest_cell.value = "=GEOMEAN({}:{})".format(start_cell.coordinate, end_cell.coordinate)

        row_offset += 1
        col_offset += 1

    cell_range.expand(down=row_offset, right=col_offset - 1)

    return cell_range


def value_ref(cell):
    return "{}!{}".format(cell.parent.title, cell.coordinate)


def ratio_ref(num_cell, denom_cell):
    return "{num} / {denom}".format(num=value_ref(num_cell), denom=value_ref(denom_cell))


def safe_ratio_ref(num_cell, denom_cell):
    return 'IF(OR({denom} = 0,{denom}=-1),"",{num} / {denom})'.format(
        num=value_ref(num_cell), denom=value_ref(denom_cell)
    )


def link_sheet_header(dest_sheet, ref_sheet, row, values=None):
    # Copy header
    dest_col = 1
    for col in range(1, ref_sheet.max_column + 1):

        ref_cell = ref_sheet.cell(row=1, column=col)

        if values != None and ref_cell.value.strip() not in values:
            continue

        dest_cell = dest_sheet.cell(row=row, column=dest_col)

        dest_cell.value = "={}".format(value_ref(ref_cell))

        dest_col += 1


def dataframe_to_sheet(wb, df, sheet_name):

    # Add to sheet
    ws = wb.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    return ws


def safe_sheet_title(raw_title):

    # Keep only file name and drop file path
    safe_title = raw_title.split("/")[-1]

    if len(safe_title) > 31:
        safe_title = safe_title[-31:]

    # Remove all non-alphanumerical characters
    regex = re.compile(r"^\W+")
    safe_title = regex.sub("", safe_title)

    # Remove bad characters from title
    for bad_char in ["-"]:
        safe_title = safe_title.replace(bad_char, "_")

    # Remove all leading zeros from title name
    safe_title = safe_title.lstrip("0")

    return safe_title


def get_task_result_files(task_name):

    task_path = os.path.join("..", "tasks", task_name)
    if not os.path.isdir(task_path):
        raise RuntimeError("Task is not found at {}".format(task_path))

    task_result_files = []
    task_result_files.append(os.path.join("..", "tasks", task_name, "config", "golden_results.txt"))
    task_result_files.append(os.path.join("..", "tasks", task_name, "latest", "parse_results.txt"))

    if not os.path.isfile(task_result_files[0]):
        raise RuntimeError(
            'Golden results of task "{}" doesn\'t exist'.format(task_result_files[0])
        )

    if not os.path.isfile(task_result_files[1]):
        raise RuntimeError(
            'Latest parse results of task "{}" doesn\'t exist'.format(task_result_files[1])
        )

    return task_result_files


if __name__ == "__main__":
    main()
